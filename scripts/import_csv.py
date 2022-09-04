# from hs_code.models import TariffDetail, Chapter
# import csv

from openpyxl import load_workbook
from music.models import (
    TariffDetail,Item
)
def import_data(filename):
    wb = load_workbook(filename=filename)
    for worksheet in wb.sheetnames:
        if worksheet == "TariffDetail":
            import_tariffs(wb["TariffDetail"])
        elif worksheet == "Item":
            import_items(wb["Item"])

def import_tariffs(worksheet):
    #TariffDetail.objects.all().delete()
    for row in worksheet.iter_rows(min_row=2):
        try:
            if row[7].value == None:
                description = ""
            else:
                description = row[8].value
            if row[6].value == None:
                subheadingCode = ""
            else:
                subheadingCode = row[6].value
            try:
                tariff = TariffDetail.objects.get(id=row[0].value)
                tariff.isSubChapter = row[1].value
                tariff.isHeading = row[2].value
                tariff.isSubHeading = row[3].value
                tariff.parent_id = row[4].value
                tariff.headingCode = row[5].value
                tariff.subheadingCode = subheadingCode
                tariff.description = description
                tariff.unitOfQty = row[8].value
                tariff.rate = row[9].value
                tariff.deleted = row[10].value
                tariff.orderRank = row[11].value
                tariff.save()
            except TariffDetail.DoesNotExist as e:
                if row[0].value != None:
                    print("Creating new Tariff Detail")
                    tariff = TariffDetail(
                        id=row[0].value,
                        isSubChapter=row[1].value,
                        isHeading=row[2].value,
                        isSubHeading=row[3].value,
                        parent_id=row[4].value,
                        headingCode=row[5].value,
                        subheadingCode=subheadingCode,
                        description=description,
                        unitOfQty=row[8].value,
                        rate=row[9].value,
                        deleted=row[10].value,
                        orderRank=row[11].value,
                    )
                    tariff.save()
        except (TariffDetail.DoesNotExist) as e:
            print(e)

def import_items(worksheet):
    id_counter = 0
    for row in worksheet.iter_rows(min_row=2):
        id_counter += 1
        tariff = TariffDetail.objects.filter(subheadingCode=row[2].value).first()

        if tariff == None:
            tariff = TariffDetail.objects.filter(subheadingCode="0402.21.90").first()

        if row[1].value == None:
            description = ""
        else:
            description = row[1].value
        if row[4].value == None:
            consideration = ""
        else:
            consideration = row[4].value
        if row[5].value == None:
            remarks = ""
        else:
            remarks = row[5].value
        try:
            item = Item.objects.get(name=row[0].value)
            item.description = description
            item.orderRank = item.id * 10
            item.tariff_detail = tariff
            item.appliedGir = row[3].value
            item.considerationStep = consideration
            item.remarks = remarks
            item.save()
        except Item.DoesNotExist as e:
            if row[0].value != None:
                print("Creating new Item")
                print(row)
                item = Item(
                    name=row[0].value,
                    description=description,
                    tariff_detail=tariff,
                    appliedGir=row[3].value,
                    considerationStep=consideration,
                    remarks=remarks,
                    deleted=False,
                )
                item.orderRank = id_counter * 10
                item.save()
# def run():
#     with open('newtariffs.csv') as file:
#         reader = csv.reader(file)
#         next(reader)  # Advance past the header
#         TariffDetail.objects.all().delete()
#         chapters = {chapter.code: chapter for chapter in Chapter.objects.all()}
#         for row in reader:
#             print(row)
#             #print(chapters)
#             chapter_code= row[1]
#             chapter = chapters.get(chapter_code)
#             tariff_detail = TariffDetail(
#                         chapter=chapter,
#                         isSubChapter=row[2],
#                         isHeading=row[3],
#                         isSubHeading=row[4],
#                         parent_id=row[5],
#                         headingCode=row[6],
#                         subheadingCode=row[7],
#                         description=row[8],
#                         unitOfQty=row[9],
#                         rate=row[10],
#                         deleted=row[11],
#                         orderRank=row[12])
#             #tariff_detail.id = 10001
#             tariff_detail.save()


def run():
    import_data("import_data.xlsx")